from __future__ import annotations

import json
import math
from pathlib import Path

import polars as pl
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TRANSFER = ROOT / "step90g_transfer"
REPORT_DIR = ROOT / "data" / "reports" / "backtests"

MARKET_COEFFICIENT = 4.980172
DEF_EPA_COEFFICIENT = 1.044827
INTERCEPT = -2.514766
RESIDUAL_CAP = 0.0425

BOOKS = ("betmgm", "fanduel", "draftkings")

TEAM_MAP = {
    "Arizona Cardinals": "ARI",
    "Atlanta Falcons": "ATL",
    "Baltimore Ravens": "BAL",
    "Buffalo Bills": "BUF",
    "Carolina Panthers": "CAR",
    "Chicago Bears": "CHI",
    "Cincinnati Bengals": "CIN",
    "Cleveland Browns": "CLE",
    "Dallas Cowboys": "DAL",
    "Denver Broncos": "DEN",
    "Detroit Lions": "DET",
    "Green Bay Packers": "GB",
    "Houston Texans": "HOU",
    "Indianapolis Colts": "IND",
    "Jacksonville Jaguars": "JAX",
    "Kansas City Chiefs": "KC",
    "Las Vegas Raiders": "LV",
    "Los Angeles Chargers": "LAC",
    "Los Angeles Rams": "LA",
    "Miami Dolphins": "MIA",
    "Minnesota Vikings": "MIN",
    "New England Patriots": "NE",
    "New Orleans Saints": "NO",
    "New York Giants": "NYG",
    "New York Jets": "NYJ",
    "Philadelphia Eagles": "PHI",
    "Pittsburgh Steelers": "PIT",
    "San Francisco 49ers": "SF",
    "Seattle Seahawks": "SEA",
    "Tampa Bay Buccaneers": "TB",
    "Tennessee Titans": "TEN",
    "Washington Commanders": "WAS",
}


def american_to_probability(price: float) -> float:
    if not math.isfinite(price) or price == 0:
        raise ValueError("Invalid American price")
    return 100.0 / (price + 100.0) if price > 0 else -price / (-price + 100.0)


def no_vig_home_probability(home: float, away: float) -> float:
    hp = american_to_probability(home)
    ap = american_to_probability(away)
    total = hp + ap
    if total <= 0:
        raise ValueError("Invalid vig total")
    return hp / total


def sigmoid(value: float) -> float:
    if value >= 0:
        z = math.exp(-value)
        return 1.0 / (1.0 + z)
    z = math.exp(value)
    return z / (1.0 + z)


def frozen_probability(market: float, def_epa: float) -> float:
    raw = sigmoid(
        INTERCEPT
        + MARKET_COEFFICIENT * market
        + DEF_EPA_COEFFICIENT * def_epa
    )
    lower = max(0.0, market - RESIDUAL_CAP)
    upper = min(1.0, market + RESIDUAL_CAP)
    return min(max(raw, lower), upper)


def log_loss(prob: float, home_win: bool) -> float:
    eps = 1e-15
    p = min(max(prob, eps), 1.0 - eps)
    return -math.log(p if home_win else 1.0 - p)


def brier(prob: float, home_win: bool) -> float:
    return (prob - (1.0 if home_win else 0.0)) ** 2


def metrics(rows: list[dict], probability_key: str) -> dict:
    usable = [
        r for r in rows
        if r["outcome"] in ("HOME", "AWAY")
        and r[probability_key] is not None
    ]

    if not usable:
        return {
            "n": 0,
            "accuracy": None,
            "brier": None,
            "log_loss": None,
        }

    correct = 0
    briers = []
    losses = []

    for r in usable:
        home_win = r["outcome"] == "HOME"
        p = r[probability_key]
        predicted_home = p >= 0.5

        if predicted_home == home_win:
            correct += 1

        briers.append(brier(p, home_win))
        losses.append(log_loss(p, home_win))

    return {
        "n": len(usable),
        "accuracy": correct / len(usable),
        "brier": sum(briers) / len(briers),
        "log_loss": sum(losses) / len(losses),
    }


def main() -> None:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    workbook_path = TRANSFER / "nfl2025_complete_validation.xlsx"
    feature_path = TRANSFER / "recent_form_features_2025.parquet"
    legacy_prediction_path = TRANSFER / "predictions_2025.parquet"

    wb = load_workbook(
        workbook_path,
        read_only=True,
        data_only=True,
    )
    ws = wb["NFL 2025"]

    values = list(ws.iter_rows(values_only=True))
    headers = list(values[0])
    source_rows = [dict(zip(headers, row)) for row in values[1:]]

    regular = [
        r for r in source_rows
        if isinstance(r["week"], int) and 1 <= r["week"] <= 18
    ]

    features = pl.read_parquet(feature_path).select(
        [
            "game_id",
            "season",
            "week",
            "home_team",
            "away_team",
            "def_epa_trend_advantage",
        ]
    )

    feature_rows = {
        row["game_id"]: row
        for row in features.to_dicts()
    }

    legacy = pl.read_parquet(legacy_prediction_path)

    legacy_rows = {
        row["game_id"]: row
        for row in legacy.select(
            [
                "game_id",
                "home_win_probability",
                "away_win_probability",
                "model_version",
            ]
        ).to_dicts()
    }

    diagnostics = []
    exclusions = []

    for source in regular:
        week = int(source["week"])

        away_name = source["away_team"]
        home_name = source["home_team"]

        away = TEAM_MAP.get(away_name)
        home = TEAM_MAP.get(home_name)

        if away is None or home is None:
            exclusions.append(
                {
                    "week": week,
                    "away_team": away_name,
                    "home_team": home_name,
                    "reason": "TEAM_MAPPING_FAILURE",
                }
            )
            continue

        game_id = f"2025_{week:02d}_{away}_{home}"

        feature = feature_rows.get(game_id)

        if feature is None:
            exclusions.append(
                {
                    "week": week,
                    "game_id": game_id,
                    "away_team": away,
                    "home_team": home,
                    "reason": "MISSING_DEF_EPA_FEATURE",
                }
            )
            continue

        book_probs = {}

        try:
            for book in BOOKS:
                away_ml = source[f"{book}_away_ml"]
                home_ml = source[f"{book}_home_ml"]

                if away_ml is None or home_ml is None:
                    raise ValueError(f"{book}: missing moneyline")

                book_probs[book] = no_vig_home_probability(
                    float(home_ml),
                    float(away_ml),
                )
        except (KeyError, TypeError, ValueError) as exc:
            exclusions.append(
                {
                    "week": week,
                    "game_id": game_id,
                    "away_team": away,
                    "home_team": home,
                    "reason": "INVALID_CORE_THREE_MONEYLINE",
                    "detail": str(exc),
                }
            )
            continue

        market = sum(book_probs.values()) / 3.0
        def_epa = feature["def_epa_trend_advantage"]

        if def_epa is None or not math.isfinite(float(def_epa)):
            exclusions.append(
                {
                    "week": week,
                    "game_id": game_id,
                    "away_team": away,
                    "home_team": home,
                    "reason": "INVALID_DEF_EPA",
                }
            )
            continue

        candidate = frozen_probability(market, float(def_epa))

        away_score = source["away_score"]
        home_score = source["home_score"]

        if away_score is None or home_score is None:
            outcome = "UNKNOWN"
        elif home_score > away_score:
            outcome = "HOME"
        elif away_score > home_score:
            outcome = "AWAY"
        else:
            outcome = "TIE"

        legacy_row = legacy_rows.get(game_id)

        diagnostics.append(
            {
                "game_id": game_id,
                "week": week,
                "away_team": away,
                "home_team": home,
                "away_score": away_score,
                "home_score": home_score,
                "outcome": outcome,
                "betmgm_home_fair": book_probs["betmgm"],
                "fanduel_home_fair": book_probs["fanduel"],
                "draftkings_home_fair": book_probs["draftkings"],
                "market_home_probability": market,
                "def_epa_trend_advantage": float(def_epa),
                "candidate_home_probability": candidate,
                "candidate_predicted_winner": (
                    home if candidate >= 0.5 else away
                ),
                "market_predicted_winner": (
                    home if market >= 0.5 else away
                ),
                "legacy_v2_home_probability": (
                    legacy_row["home_win_probability"]
                    if legacy_row else None
                ),
            }
        )

    candidate_metrics = metrics(
        diagnostics,
        "candidate_home_probability",
    )

    market_metrics = metrics(
        diagnostics,
        "market_home_probability",
    )

    v2_metrics = metrics(
        diagnostics,
        "legacy_v2_home_probability",
    )

    ties = sum(r["outcome"] == "TIE" for r in diagnostics)

    week_metrics = {}

    for week in range(1, 19):
        subset = [r for r in diagnostics if r["week"] == week]
        week_metrics[str(week)] = {
            "games": len(subset),
            "ties": sum(r["outcome"] == "TIE" for r in subset),
            "candidate": metrics(subset, "candidate_home_probability"),
            "market": metrics(subset, "market_home_probability"),
        }

    report = {
        "classification": "HISTORICAL_DIAGNOSTIC_NOT_PROSPECTIVE_EVIDENCE",
        "protocol": "Step 91O Phase 4B",
        "season": 2025,
        "population": "REG Weeks 1-18",
        "source_games": len(source_rows),
        "regular_season_games": len(regular),
        "diagnostic_rows": len(diagnostics),
        "excluded_rows": len(exclusions),
        "ties": ties,
        "candidate": {
            "id": "market-plus-def-epa-capped-0425-v1",
            "market_coefficient": MARKET_COEFFICIENT,
            "def_epa_coefficient": DEF_EPA_COEFFICIENT,
            "intercept": INTERCEPT,
            "residual_cap": RESIDUAL_CAP,
            "market_universe": list(BOOKS),
            "weight_per_book": 1.0 / 3.0,
        },
        "results": {
            "frozen_candidate": candidate_metrics,
            "three_book_market_only": market_metrics,
            "legacy_v2_reference": v2_metrics,
        },
        "weekly": week_metrics,
        "exclusions": exclusions,
        "limitations": [
            "Historical sportsbook timestamps and feed identity are not preserved.",
            "Historical Core-Three prices are not a conformed prospective atomic provider response.",
            "Historical EPA source-vintage provenance is unresolved.",
            "No fitting, recalibration, optimization, or coefficient changes were performed.",
            "Results are diagnostic only and are not prospective evidence.",
        ],
    }

    json_path = REPORT_DIR / "phase4b_2025_frozen_core_three_diagnostic.json"
    md_path = REPORT_DIR / "phase4b_2025_frozen_core_three_diagnostic.md"
    csv_path = REPORT_DIR / "phase4b_2025_frozen_core_three_diagnostic_rows.csv"

    json_path.write_text(
        json.dumps(report, indent=2, sort_keys=True),
        encoding="utf-8",
    )

    pl.DataFrame(diagnostics).write_csv(csv_path)

    lines = [
        "# Step 91O Phase 4B — 2025 Frozen Core-Three Historical Diagnostic",
        "",
        "**Classification: HISTORICAL DIAGNOSTIC — NOT PROSPECTIVE EVIDENCE**",
        "",
        "## Population",
        "",
        f"- Source rows: {len(source_rows)}",
        f"- 2025 regular-season games: {len(regular)}",
        f"- Diagnostic rows: {len(diagnostics)}",
        f"- Excluded rows: {len(exclusions)}",
        f"- Ties: {ties}",
        "",
        "## Frozen Candidate",
        "",
        "- Market coefficient: 4.980172",
        "- DEF EPA coefficient: 1.044827",
        "- Intercept: -2.514766",
        "- Residual cap: ±4.25%",
        "- Market books: BetMGM / FanDuel / DraftKings",
        "- Weight: 1/3 each",
        "",
        "## Results",
        "",
        "| Model | N | Accuracy | Brier | Log loss |",
        "|---|---:|---:|---:|---:|",
    ]

    for label, result in (
        ("Frozen candidate", candidate_metrics),
        ("Core-Three market only", market_metrics),
        ("Legacy v2 reference", v2_metrics),
    ):
        lines.append(
            f"| {label} | {result['n']} | "
            f"{result['accuracy']:.4f} | "
            f"{result['brier']:.4f} | "
            f"{result['log_loss']:.4f} |"
        )

    lines.extend(
        [
            "",
            "## Interpretation Boundary",
            "",
            "This replay applies the frozen candidate mechanically. "
            "No model was fitted or recalibrated.",
            "",
            "The historical market and EPA provenance limitations identified "
            "by Step 91O Phase 4A remain unresolved. Therefore these results "
            "must not be treated as leakage-safe validation or prospective evidence.",
            "",
            "## Files",
            "",
            f"- `{json_path.relative_to(ROOT)}`",
            f"- `{md_path.relative_to(ROOT)}`",
            f"- `{csv_path.relative_to(ROOT)}`",
        ]
    )

    md_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    print("=" * 88)
    print("STEP 91O PHASE 4B — FROZEN CORE-THREE HISTORICAL DIAGNOSTIC")
    print("=" * 88)
    print(f"Regular-season games: {len(regular)}")
    print(f"Diagnostic rows:      {len(diagnostics)}")
    print(f"Excluded:             {len(exclusions)}")
    print(f"Ties:                 {ties}")
    print()
    print("MODEL                         N       ACC      BRIER    LOG LOSS")
    print("-" * 72)

    for label, result in (
        ("Frozen candidate", candidate_metrics),
        ("Core-Three market", market_metrics),
        ("Legacy v2", v2_metrics),
    ):
        print(
            f"{label:<28} "
            f"{result['n']:>3} "
            f"{result['accuracy']:>9.4f} "
            f"{result['brier']:>9.4f} "
            f"{result['log_loss']:>10.4f}"
        )

    print()
    print(f"Wrote: {json_path}")
    print(f"Wrote: {md_path}")
    print(f"Wrote: {csv_path}")


if __name__ == "__main__":
    main()
